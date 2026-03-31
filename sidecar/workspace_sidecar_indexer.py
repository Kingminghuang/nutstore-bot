from __future__ import annotations

import csv
from datetime import date, datetime, time, timezone
import json
import logging
import os
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET
import zipfile

from fastapi import BackgroundTasks


LOGGER = logging.getLogger(__name__)

SUPPORTED_SOURCE_EXTENSIONS = {".pdf", ".docx", ".xlsx"}
MARKDOWN_SOURCE_EXTENSIONS = {".pdf", ".docx"}
MANIFEST_FILE_NAME = ".index-manifest.json"


class WorkspaceSidecarIndexer:
    def enqueue(
        self,
        background_tasks: BackgroundTasks | None,
        workspace_id: str,
        workspace_real_path: str,
    ) -> None:
        if background_tasks is None:
            self._run_indexing_task(workspace_id, workspace_real_path)
            return
        background_tasks.add_task(
            self._run_indexing_task,
            workspace_id,
            workspace_real_path,
        )

    def _run_indexing_task(self, workspace_id: str, workspace_real_path: str) -> None:
        try:
            self.index_workspace(workspace_id, workspace_real_path)
        except Exception:  # noqa: BLE001
            LOGGER.exception(
                "Workspace sidecar indexing failed",
                extra={
                    "workspace_id": workspace_id,
                    "workspace_path": workspace_real_path,
                },
            )

    def index_workspace(self, workspace_id: str, workspace_real_path: str) -> None:
        workspace_path = Path(workspace_real_path).expanduser().resolve()
        if not workspace_path.exists() or not workspace_path.is_dir():
            LOGGER.warning(
                "Skip workspace sidecar indexing because workspace is missing",
                extra={
                    "workspace_id": workspace_id,
                    "workspace_path": str(workspace_path),
                },
            )
            return

        sidecar_root = workspace_path / ".sidecar"
        sidecar_root.mkdir(parents=True, exist_ok=True)
        manifest_path = sidecar_root / MANIFEST_FILE_NAME
        manifest = _load_manifest(manifest_path)
        sources = manifest.get("sources", {})
        if not isinstance(sources, dict):
            sources = {}

        last_indexed_at = datetime.now(timezone.utc).isoformat()
        scanned_count = 0
        converted_count = 0
        skipped_count = 0
        failed_count = 0

        for source_file in _iter_supported_files(workspace_path):
            scanned_count += 1
            relative_path = source_file.relative_to(workspace_path)
            relative_key = relative_path.as_posix()
            output_file = _build_output_path(sidecar_root, relative_path)
            signature = _build_signature(source_file)
            manifest_entry = sources.get(relative_key)

            if _is_unchanged(manifest_entry, signature, output_file):
                skipped_count += 1
                continue

            output_file.parent.mkdir(parents=True, exist_ok=True)
            try:
                _convert_file(source_file, output_file)
            except Exception:  # noqa: BLE001
                failed_count += 1
                LOGGER.exception(
                    "Workspace sidecar file conversion failed",
                    extra={
                        "workspace_id": workspace_id,
                        "workspace_path": str(workspace_path),
                        "source_file": str(source_file),
                        "output_file": str(output_file),
                    },
                )
                continue

            sources[relative_key] = {
                "signature": signature,
                "output": output_file.relative_to(sidecar_root).as_posix(),
                "updatedAt": last_indexed_at,
            }
            converted_count += 1

        manifest["workspaceId"] = workspace_id
        manifest["workspacePath"] = str(workspace_path)
        manifest["lastIndexedAt"] = last_indexed_at
        manifest["stats"] = {
            "scanned": scanned_count,
            "converted": converted_count,
            "skipped": skipped_count,
            "failed": failed_count,
        }
        manifest["sources"] = sources

        _write_manifest(manifest_path, manifest)
        LOGGER.info(
            "Workspace sidecar indexing completed",
            extra={
                "workspace_id": workspace_id,
                "workspace_path": str(workspace_path),
                "scanned": scanned_count,
                "converted": converted_count,
                "skipped": skipped_count,
                "failed": failed_count,
            },
        )

    def status(self, workspace_id: str, workspace_real_path: str) -> dict[str, Any]:
        workspace_path = Path(workspace_real_path).expanduser().resolve()
        sidecar_root = workspace_path / ".sidecar"
        manifest_path = sidecar_root / MANIFEST_FILE_NAME
        payload = _load_manifest(manifest_path)
        sources_payload = payload.get("sources", {})
        sources = sources_payload if isinstance(sources_payload, dict) else {}

        last_indexed_at = payload.get("lastIndexedAt")
        stats_payload = payload.get("stats", {})
        stats = stats_payload if isinstance(stats_payload, dict) else {}

        status = "indexed" if manifest_path.exists() and last_indexed_at else "not_started"
        return {
            "workspaceId": workspace_id,
            "workspacePath": str(workspace_path),
            "sidecarRoot": str(sidecar_root),
            "manifestPath": str(manifest_path),
            "manifestExists": manifest_path.exists(),
            "status": status,
            "lastIndexedAt": last_indexed_at,
            "stats": {
                "scanned": int(stats.get("scanned", 0)),
                "converted": int(stats.get("converted", 0)),
                "skipped": int(stats.get("skipped", 0)),
                "failed": int(stats.get("failed", 0)),
            },
            "sourceCount": len(sources),
        }


def _iter_supported_files(workspace_path: Path):
    for root, dir_names, file_names in os.walk(workspace_path, topdown=True):
        dir_names[:] = [name for name in dir_names if not name.startswith(".")]
        for file_name in file_names:
            source_path = Path(root) / file_name
            if source_path.suffix.lower() in SUPPORTED_SOURCE_EXTENSIONS:
                yield source_path


def _build_output_path(sidecar_root: Path, source_relative_path: Path) -> Path:
    base_output = sidecar_root / source_relative_path
    if source_relative_path.suffix.lower() in MARKDOWN_SOURCE_EXTENSIONS:
        return base_output.with_name(base_output.name + ".md")
    return base_output.with_name(base_output.name + ".csv")


def _build_signature(source_path: Path) -> dict[str, int]:
    stat_result = source_path.stat()
    return {"mtimeNs": stat_result.st_mtime_ns, "size": stat_result.st_size}


def _is_unchanged(
    manifest_entry: Any, signature: dict[str, int], output_path: Path
) -> bool:
    if not isinstance(manifest_entry, dict):
        return False
    return manifest_entry.get("signature") == signature and output_path.exists()


def _convert_file(source_path: Path, output_path: Path) -> None:
    suffix = source_path.suffix.lower()
    if suffix == ".pdf":
        _convert_pdf_to_markdown(source_path, output_path)
        return
    if suffix == ".docx":
        _convert_docx_to_markdown(source_path, output_path)
        return
    if suffix == ".xlsx":
        _convert_xlsx_to_csv(source_path, output_path)
        return
    raise ValueError(f"Unsupported file extension: {source_path.suffix}")


def _convert_pdf_to_markdown(source_path: Path, output_path: Path) -> None:
    from pypdf import PdfReader

    reader = PdfReader(str(source_path))
    page_chunks: list[str] = []
    for index, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if not text:
            text = "(empty page)"
        page_chunks.append(f"## Page {index}\n\n{text}")

    output_path.write_text("\n\n---\n\n".join(page_chunks).strip() + "\n", encoding="utf-8")


def _convert_docx_to_markdown(source_path: Path, output_path: Path) -> None:
    markdown: str | None = None

    try:
        import mammoth
        from markdownify import markdownify as html_to_markdown

        with source_path.open("rb") as source_file:
            html_result = mammoth.convert_to_html(source_file)
        markdown = html_to_markdown(html_result.value or "", heading_style="ATX")
    except Exception:  # noqa: BLE001
        LOGGER.warning(
            "Falling back to plain-text DOCX conversion",
            extra={"source_file": str(source_path)},
        )

    if markdown is None or markdown.strip() == "":
        markdown = _extract_docx_plain_text(source_path)

    normalized = markdown.strip() if markdown else ""
    output_path.write_text((normalized + "\n") if normalized else "", encoding="utf-8")


def _extract_docx_plain_text(source_path: Path) -> str:
    xml_bytes: bytes
    with zipfile.ZipFile(source_path) as archive:
        xml_bytes = archive.read("word/document.xml")

    root = ET.fromstring(xml_bytes)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

    paragraphs: list[str] = []
    for paragraph in root.findall(".//w:p", namespace):
        fragments = [
            node.text or "" for node in paragraph.findall(".//w:t", namespace)
        ]
        line = "".join(fragments).strip()
        if line:
            paragraphs.append(line)

    if paragraphs:
        return "\n\n".join(paragraphs)

    all_text = [node.text or "" for node in root.findall(".//w:t", namespace)]
    return "\n".join(part for part in all_text if part).strip()


def _convert_xlsx_to_csv(source_path: Path, output_path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=str(source_path), read_only=True, data_only=True)
    sheet_rows: list[tuple[str, list[list[Any]]]] = []
    max_columns = 0
    try:
        for sheet in workbook.worksheets:
            rows: list[list[Any]] = []
            for row in sheet.iter_rows(values_only=True):
                row_values = list(row)
                max_columns = max(max_columns, len(row_values))
                rows.append(row_values)
            sheet_rows.append((sheet.title, rows))
    finally:
        workbook.close()

    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.writer(output_file)
        writer.writerow(["__sheet", *[f"col_{index}" for index in range(1, max_columns + 1)]])
        for sheet_name, rows in sheet_rows:
            for row_values in rows:
                padded_row = row_values + [None] * (max_columns - len(row_values))
                writer.writerow(
                    [sheet_name, *[_to_csv_value(value) for value in padded_row]]
                )


def _to_csv_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return value


def _load_manifest(manifest_path: Path) -> dict[str, Any]:
    if not manifest_path.exists():
        return {"sources": {}}
    try:
        loaded = json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001
        LOGGER.warning("Unable to parse existing sidecar manifest; rebuilding")
        return {"sources": {}}
    if not isinstance(loaded, dict):
        return {"sources": {}}
    return loaded


def _write_manifest(manifest_path: Path, payload: dict[str, Any]) -> None:
    temporary_path = manifest_path.with_suffix(f"{manifest_path.suffix}.tmp")
    temporary_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    temporary_path.replace(manifest_path)
